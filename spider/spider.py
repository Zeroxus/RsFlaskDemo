import requests
from openpyxl import Workbook,load_workbook
from bs4 import BeautifulSoup
from config import *


class spider:
    def __init__(self,url,cookie,session):
        self.url = url
        self.cookie = cookie
        self.session = session

    def makeRequest(self,pageNum):
        s = self.session
        url = self.getPage(pageNum)
        req = s.get(url=url, cookies=cookie, headers=headers, allow_redirects=False)
        return req

    def makeData(self,pageNum):
        targetpage = self.makeRequest(pageNum)
        soup = BeautifulSoup(targetpage.content,'lxml')
        data_list=[]
        for idx,tr in enumerate(soup.find_all('tr')):
            if idx!=0:
                tds = tr.find_all('td')
                data_list.append([
                    tds[0].contents[0],
                    tds[1].contents[0],
                    ''.join(tds[2].contents[0].contents),
                    tds[3].contents[0],
                    tds[4].contents[0],
                    tds[5].contents[0],
                    tds[6].contents[0],
                ])
        return data_list

    def getPage(self,pageNum):
        spider_url = self.url+str(pageNum)
        print(spider_url)
        return spider_url

    def createExcel(self):
        wb = Workbook()
        ws = wb.create_sheet(title='书籍列表')
        row0 = ['编号','条码号','题名','责任者','借阅日期','归还日期','馆藏地']
        ws.append(row0)
        save_path = 'book_hist.xlsx'
        wb.save(save_path)

    def saveToExcel(self,data_list):
        wb = load_workbook(filename='book_hist.xlsx')
        ws = wb.get_sheet_by_name('书籍列表')
        for i in range(len(data_list)):
            ws.append(data_list[i])
        save_path = 'book_hist.xlsx'
        wb.save(save_path)


class login:
    def __init__(self):
        pass

    def getCookie(self):
        pass

    def getHeaders(self):
        pass

    def getXsrf(self):
        pass

    def getCaptcha(self):
        pass



if __name__=='__main__':
    session = requests.Session()
    loginResponse=session.post(url=loginUrl,headers=headers,data=postData)
    cookie = loginResponse.cookies
    s = spider(targetUrl,cookie,session)
    s.createExcel()
    for i in range(1,8):
        book_list = s.makeData(i)
        s.saveToExcel(book_list)